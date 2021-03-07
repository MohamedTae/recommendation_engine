# -*- coding: utf-8 -*-
import requests
import openpyxl
import wikipedia
from imdb import IMDb
import re
from geotext import GeoText
import urllib2


ia = IMDb()
wikipedia.set_lang("en")

payload = {}
headers = {}

wb = openpyxl.load_workbook(filename='staff_lighttiger.xlsx', data_only=True)
a_sheet_names = wb.get_sheet_names()
o_sheet = wb.get_sheet_by_name("staff_lighttiger")

for x in range(100,10000) :
    try:
        o_cell = o_sheet.cell(row=x, column=2)
        #print(o_cell.value)
        dd = ia.search_person(o_cell.value)
        ss = str(dd)
        pattern = "id:(.*?)\["
        substring = re.search(pattern, ss).group(1)
        #print(substring)
        url = "https://www.imdb.com/name/nm"+str(substring)+"/bio?ref_=nm_ov_bio_sm"
        response = requests.request("GET", url, headers=headers, data=payload)
        o = str(response.text.encode('utf8'))
        substring_2 = re.search('place=(.*?)"', o).group(1)
        pp = urllib2.unquote(substring_2)
        #print(pp)
        o_sheet.cell(row=x, column=3).value = pp

    except:
        try:
            o_cell = o_sheet.cell(row=x, column=2)
            print(o_cell.value)
            d = wikipedia.page(o_cell.value,auto_suggest=True,redirect=True).content
            places = GeoText(d)
            e = places.nationalities[0]
            o_sheet.cell(row=x, column=3).value = e
            print(e)
        except:
            print("Not Found")

wb.save('staff_lighttiger_1.xlsx')
